import argparse
from collections import defaultdict
import json
import os
import sys
import glob
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from src.models import get_dataset_spec
from src.mechanisms.topk import topk_sparsify, update_size_bytes


def _io_path(path: str) -> str:
    """Return an absolute path suitable for filesystem I/O on this OS."""
    absolute_path = os.path.abspath(path)
    if os.name != "nt":
        return absolute_path

    if absolute_path.startswith("\\\\?\\") or len(absolute_path) < 260:
        return absolute_path

    if absolute_path.startswith("\\\\"):
        return "\\\\?\\UNC\\" + absolute_path[2:]
    return "\\\\?\\" + absolute_path


def load_results(folder: str) -> list[dict]:
    """
    Load all JSON result files from a folder.

    Args:
        folder (str): path to folder containing result JSON files.

    Returns:
        list[dict]: list of loaded result dicts, sorted by filename.
    """
    pattern = os.path.join(folder, "*.json")
    files = sorted(
        f for f in glob.glob(pattern)
        if os.path.basename(f) != "run_summary.json"
    )

    if not files:
        print(f"No JSON files found in {folder}")
        sys.exit(1)

    results = []
    for filepath in files:
        with open(_io_path(filepath)) as f:
            data = json.load(f)
            data["_filename"] = os.path.basename(filepath)
            results.append(data)

    print(f"Loaded {len(results)} result files from {folder}")
    return results


def load_results_if_present(folder: str) -> list[dict]:
    """
    Like load_results(), but returns [] instead of exiting when folder
    doesn't exist or contains no JSON result files.

    Used for optional subfolders (e.g. clean_base_run/, see
    export_excel_report()) that aren't present under every results folder.

    Args:
        folder (str): path to folder containing result JSON files.

    Returns:
        list[dict]: list of loaded result dicts, sorted by filename, or [].
    """
    if not os.path.isdir(folder):
        return []
    pattern = os.path.join(folder, "*.json")
    files = sorted(
        f for f in glob.glob(pattern)
        if os.path.basename(f) != "run_summary.json"
    )

    results = []
    for filepath in files:
        with open(_io_path(filepath)) as f:
            data = json.load(f)
            data["_filename"] = os.path.basename(filepath)
            results.append(data)
    return results


def make_label(filename: str) -> str:
    """
    Generate a short human-readable label from a result filename.

    Extracts dp, fltrust, topk fields from the filename.

    Args:
        filename (str): result JSON filename.

    Returns:
        str: short label e.g. 'dp-False_fltrust-True_topk-False'
    """
    parts  = filename.replace(".json", "").split("_")
    keep   = ["dp", "fltrust", "topk"]
    labels = []
    i = 0
    while i < len(parts):
        if parts[i].startswith("config-"):
            labels.append(replace_config_with_label(parts[i]))
        for key in keep:
            if parts[i].startswith(key):
                # include epsilon if dp=True
                if key == "dp" and parts[i] == "dp-True" and i + 1 < len(parts):
                    if parts[i + 1].startswith("epsilon"):
                        labels.append(f"{parts[i]}_{parts[i+1]}")
                        i += 1
                        break
                # include k if topk=True
                if key == "topk" and parts[i] == "topk-True" and i + 1 < len(parts):
                    if parts[i + 1].startswith("k"):
                        labels.append(f"{parts[i]}_{parts[i+1]}")
                        i += 1
                        break
                labels.append(parts[i])
                break
        i += 1
    return "_".join(labels)

def replace_config_with_label(config_part: str) -> str:
    """
    Replace 'config-X' with a more descriptive label based on the experimental grid.

    Args:
        config_part (str): part of the filename starting with 'config-'.

    Returns:
        str: descriptive label for the config.
    """
    config_id = config_part.replace("config-", "")
    config_labels = {
        "1": "BASE",
        "2": "DP",
        "3": "FLTrust",
        "4": "TopK",
        "5": "DP + FLTrust",
        "6": "DP + TopK",
        "7": "FLTrust + TopK",
        "8": "ALL"
    }
    return config_labels.get(config_id, f"Config-{config_id}")

def variant_sort_key(result: dict) -> tuple:
    """
    Sort key ordering same-config variants ascending by their swept parameter(s).

    Reads the actual numeric config fields rather than the filename string --
    sorting by filename text breaks as soon as a swept value has a different
    digit count (e.g. epsilon 5.0 would sort after 10.0, since "5" > "1"
    lexicographically). Priority order (outermost/slowest-changing first):
    topk_ratio, epsilon, num_client_iterations_per_round, num_clients --
    matches tmp/template_excel_output.xlsx's hand-built column layout, reused
    here for both bar_accuracy's bar ordering and the Excel table export.

    Args:
        result (dict): loaded result JSON.

    Returns:
        tuple: ascending sort key.
    """
    config = result["config"]
    return (
        config.get("topk_ratio", 0.0),
        config.get("epsilon", 0.0),
        config.get("num_client_iterations_per_round") or 0,
        config.get("num_clients", 0),
    )


def extract_metric(result: dict, metric: str) -> tuple[list[int], list[float]]:
    """
    Extract per-round values for a given metric from a result dict.

    Args:
        result (dict): loaded result JSON.
        metric (str):  one of 'loss', 'accuracy', 'epsilon'.

    Returns:
        tuple: (rounds, values) - both lists, None values filtered out.
    """
    rounds = []
    values = []
    for entry in result["results"]["per_round"]:
        val = entry.get(metric)
        if val is not None:
            rounds.append(entry["round"])
            values.append(val)
    return rounds, values


def visualize_overview(folder: str) -> None:
    """
    Load results from folder and plot loss, accuracy, epsilon per round.

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = load_results(folder)
    metrics = ["loss", "accuracy", "epsilon"]
    titles  = {
        "loss":     "Loss per Round",
        "accuracy": "Accuracy per Round",
        "epsilon":  "Cumulative Epsilon per Round (DP configs only)",
    }

    fig, axes = plt.subplots(1, 3, figsize=(18, 5))
    fig.suptitle(f"FL Trilemma Results {os.path.basename(folder)}", fontsize=14)

    for ax, metric in zip(axes, metrics):
        ax.set_title(titles[metric])
        ax.set_xlabel("Round")
        ax.set_ylabel(metric.capitalize())
        ax.grid(True, alpha=0.3)

        any_plotted = False
        for result in results:
            rounds, values = extract_metric(result, metric)
            if not values:
                continue
            label = make_label(result["_filename"])
            ax.plot(rounds, values, label=label)
            any_plotted = True

        if any_plotted:
            ax.legend(
                fontsize=6,
                loc="upper center",
                bbox_to_anchor=(0.5, -0.25),    # place below the axis
                ncol=2,                         # two columns to keep it compact
                framealpha=0.8,
            )
        else:
            ax.text(0.5, 0.5, "No data", ha="center", va="center",
                    transform=ax.transAxes, color="gray")

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.35)   # make room for legends below
    output_path = os.path.join(folder, "visualization.png")
    plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
    print(f"Saved to: {output_path}")
    plt.close()


def visualize_lines_per_config(folder: str) -> None:
    """
    One figure per config, three subplots (loss, accuracy, epsilon).
    
    One curve per variant within each config.
    Saves 8 separate PNG files, one per config.

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = load_results(folder)

    config_groups = defaultdict(list)
    for result in results:
        config_groups[result["config"]["config_id"]].append(result)

    metrics = ["loss", "accuracy", "epsilon"]
    titles  = {
        "loss":     "Loss per Round",
        "accuracy": "Accuracy per Round",
        "epsilon":  "Epsilon per Round",
    }

    # Output dir
    subfolder_name = "per_config_line_charts"
    os.makedirs(os.path.join(folder, subfolder_name), exist_ok=True)

    for config_id, group_results in sorted(config_groups.items()):
        fig, axes = plt.subplots(1, 3, figsize=(15, 5))
        fig.suptitle(replace_config_with_label(str(config_id)) + " - Per Round Metrics", fontsize=13)

        for ax, metric in zip(axes, metrics):
            ax.set_title(titles[metric])
            ax.set_xlabel("Round")
            ax.set_ylabel(metric.capitalize())
            ax.grid(True, alpha=0.3)

            for result in group_results:
                rounds, values = extract_metric(result, metric)
                if not values:
                    continue
                label = make_label(result["_filename"])
                ax.plot(rounds, values, label=label)

            # Make sure empty plots dont throw warnings if there is nothing to plot
            handles, _ = ax.get_legend_handles_labels()
            if handles:
                ax.legend(
                    fontsize=7,
                    loc="upper center",
                    bbox_to_anchor=(0.5, -0.18),
                    ncol=1,
                    framealpha=0.8,
                )
            else:
                ax.text(0.5, 0.5, "N/A for this config",
                        ha="center", va="center",
                        transform=ax.transAxes, color="gray", fontsize=9)

        plt.tight_layout()
        plt.subplots_adjust(bottom=0.25)
        output_path = os.path.join(folder, subfolder_name, f"line_chart_config_{config_id}.png")
        plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
        print(f"Saved to: {output_path}")
        plt.close()


def visualize_bar_chart_per_config(folder: str) -> None:
    """
    Bar chart of final accuracy per variant, grouped by config.

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = load_results(folder)
    
    config = results[0].get('config') if results else None
    dataset = config.get('dataset', 'Undefined') if config else 'Undefined'

    # Group results
    groups = {}
    for result in results:
        config_id = result["config"]["config_id"]
        groups.setdefault(config_id, []).append(result)

    # Build bar data
    config_ids = sorted(groups.keys())
    group_bars = []             # [(label, accuracy)] per group
    for config_id in config_ids:
        bars = []
        for result in sorted(groups[config_id], key=variant_sort_key):
            # Take only final rounds accuracy
            final_accuracy = result["results"]["per_round"][-1].get("accuracy")
            if final_accuracy is None:
                print(f"Could not find accuracy of final round for one of configs {config_id} results. ({result['_filename']})")
                continue
            label = make_label(result["_filename"])
            bars.append((label, final_accuracy))
        group_bars.append((config_id, bars))

    # Plot
    fig, ax = plt.subplots(figsize=(16, 6))
    ax.set_title(f"Final Round Accuracy per Config Variant ({dataset})", fontsize=13)
    ax.set_ylabel("Accuracy")
    ax.set_ylim(0, 1.05)
    ax.grid(axis="y", alpha=0.3)

    x         = 0          # current x position
    tick_pos  = []         # x positions for group labels
    tick_lab  = []         # group label text
    bar_width = 0.6
    gap       = 1.2        # gap between config groups

    colors = plt.cm.tab10.colors
    hatches = ['--', '++', 'xx', '//', '\\\\', '||', '*', 'o', '.', 'x']

    for config_id, bars in group_bars:
        group_start = x
        for i, (label, acc) in enumerate(bars):
            color = colors[config_id % len(colors)]
            hatch = hatches[i % len(hatches)]
            bar = ax.bar(x, acc, width=bar_width,
                         color=color, alpha=0.85,
                         hatch=hatch, edgecolor="white",
                         label=label)
            ax.text(x, acc + 0.01, f"{acc:.2f}",
                    ha="center", va="bottom", fontsize=6, rotation=45)
            x += bar_width + 0.1

        # center the group label under its bars
        group_center = (group_start + x - bar_width - 0.1) / 2
        tick_pos.append(group_center)
        tick_lab.append(replace_config_with_label(str(config_id)))
        x += gap

    ax.set_xticks(tick_pos)
    ax.set_xticklabels(tick_lab, fontsize=9)

    # Legend below plot
    handles = []
    for config_id, bars in group_bars:
        for i, (label, _) in enumerate(bars):
            handles.append(plt.Rectangle(
                (0, 0), 1, 1,
                facecolor=colors[config_id % len(colors)],
                hatch=hatches[i % len(hatches)],
                edgecolor="white",
                label=label,
                alpha=1,
            ))
    legend = None
    if handles:
        legend = ax.legend(
            handles=handles,
            fontsize=6,
            loc="upper center",
            bbox_to_anchor=(0.5, -0.12),
            ncol=4,
            framealpha=0.8,
            handleheight=2,
        )

    plt.tight_layout()
    plt.subplots_adjust(bottom=0.3)

    # Run params, positioned just below the legend's actual rendered bottom
    # edge. Computed after layout so it hugs a short legend (few variants)
    # and still clears a tall, many-row one (many variants), rather than a
    # fixed offset tuned for one particular legend size. Legend.get_window_extent()
    # needs a draw() first so the renderer has real (not stale) coordinates,
    # and Legend defaults to zorder=5 vs. Text's zorder=3, which is why a
    # fixed-offset text used to render underneath a tall legend instead of
    # just missing it.
    run_config = results[0]["config"]
    attack_type = run_config.get("attack_type", "label_flip")
    attack_scale = run_config.get("attack_scale")
    attack_text = f"Attack: {attack_type}"
    if attack_scale is not None and attack_scale != 1.0:
        attack_text += f" (scale x{attack_scale})"
    params_text = (
        f"Clients: {run_config['num_clients']}  |  "
        f"Byzantine: {run_config['num_byzantine']}  |  "
        f"Rounds: {run_config['num_rounds']}  |  "
        f"Root Dataset Size: {run_config['root_dataset_size']}  |  "
        f"{attack_text}"
    )
    if legend is not None:
        fig.canvas.draw()
        legend_bbox_axes = legend.get_window_extent(fig.canvas.get_renderer()).transformed(ax.transAxes.inverted())
        text_y = legend_bbox_axes.y0 - 0.03
    else:
        text_y = -0.05
    ax.text(0.5, text_y, params_text, ha="center", va="top",
            fontsize=8, transform=ax.transAxes)

    output_path = os.path.join(folder, "bar_accuracy.png")
    plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
    print(f"Saved to: {output_path}")
    plt.close()


def _avg_seconds_per_round(result: dict) -> float | None:
    """
    Average wall-clock duration per round (s) for a single run.

    Mirrors the "Avg s/round" figure in the Excel "Elapsed Time" sheet
    (see _write_elapsed_time_sheet()): elapsed_seconds / num_rounds, since
    a run only records total wall-clock time, not a per-round breakdown.

    Args:
        result (dict): loaded result JSON.

    Returns:
        float | None: seconds per round, or None if elapsed_seconds/num_rounds
                      is missing.
    """
    elapsed = result["results"].get("elapsed_seconds")
    num_rounds = result["config"]["num_rounds"]
    if elapsed is None or not num_rounds:
        return None
    return elapsed / num_rounds


def visualize_duration_per_config(folder: str) -> None:
    """
    Avg wall-clock duration per round (s) vs. number of clients, one
    subplot per config family, one line per variant within that config
    (e.g. DP's two epsilon values, or DP+TopK's four epsilon x topk
    combos) -- so a variant's line spans all three client counts (10, 30, 60).

    Faceted per config rather than one combined chart: pooling every
    variant across all 8 configs would put ~18 lines on a single axis
    (config x epsilon x topk combos), which is too cluttered to read.

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = load_results(folder)

    config_groups = defaultdict(list)
    for result in results:
        config_groups[result["config"]["config_id"]].append(result)

    # Group each config's variants by their non-client-count parameters
    # (epsilon / topk_ratio) so each swept-parameter combo becomes its own
    # line across client counts, rather than one line per result. Built
    # up front (rather than per-subplot) so the global max duration below
    # can be computed before any axis is drawn.
    config_series = {}
    global_max = 0.0
    for config_id, group_results in config_groups.items():
        series = defaultdict(list)
        for result in group_results:
            config = result["config"]
            series_key = (
                config.get("epsilon") if config.get("use_dp") else None,
                config.get("topk_ratio") if config.get("use_topk") else None,
            )
            series[series_key].append(result)

        lines = {}
        for series_key, variants in series.items():
            variants = sorted(variants, key=lambda r: r["config"]["num_clients"])
            clients, durations = [], []
            for result in variants:
                duration = _avg_seconds_per_round(result)
                if duration is None:
                    continue
                clients.append(result["config"]["num_clients"])
                durations.append(duration)
            if durations:
                lines[series_key] = (clients, durations)
                global_max = max(global_max, max(durations))
        config_series[config_id] = lines

    # Shared ceiling for every subplot's y-axis (with a little headroom) so
    # bar heights/slopes are visually comparable across configs.
    shared_ylim = global_max * 1.1 if global_max > 0 else 1.0

    config_ids = sorted(config_groups.keys())
    cols = min(4, len(config_ids))
    rows = (len(config_ids) + cols - 1) // cols
    fig, axes = plt.subplots(rows, cols, figsize=(cols * 4.5, rows * 4), squeeze=False)
    fig.suptitle("Avg Duration per Round vs. Number of Clients, per Config", fontsize=14)
    axes = axes.flatten()

    for idx, config_id in enumerate(config_ids):
        ax = axes[idx]
        ax.set_title(replace_config_with_label(str(config_id)), fontsize=10)
        ax.set_xlabel("Clients", fontsize=8)
        ax.set_ylabel("Avg s/round", fontsize=8)
        ax.set_ylim(0, shared_ylim)
        ax.grid(True, alpha=0.3)

        any_plotted = False
        for series_key in sorted(config_series[config_id], key=lambda k: (k[0] or 0.0, k[1] or 0.0)):
            clients, durations = config_series[config_id][series_key]
            epsilon, topk_ratio = series_key
            label_parts = []
            if epsilon is not None:
                label_parts.append(f"epsilon={epsilon}")
            if topk_ratio is not None:
                label_parts.append(f"k={topk_ratio}")
            label = ", ".join(label_parts) if label_parts else "default"

            ax.plot(clients, durations, marker="o", label=label)
            any_plotted = True

        if any_plotted:
            ax.legend(fontsize=7, loc="best")
        else:
            ax.text(0.5, 0.5, "No data", ha="center", va="center",
                    transform=ax.transAxes, color="gray", fontsize=9)

    for idx in range(len(config_ids), len(axes)):
        axes[idx].set_visible(False)

    plt.tight_layout()
    output_path = os.path.join(folder, "duration_per_config.png")
    plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
    print(f"Saved to: {output_path}")
    plt.close()


def _box_border(ws, row1, col1, row2, col2, style="thin", top=True, bottom=True, left=True, right=True):
    """
    Draw a rectangular border around a cell range, preserving each cell's
    existing border sides (openpyxl's Border is immutable, so a fresh Border
    combining old + new sides has to be rebuilt per cell rather than just
    assigning one side).

    Args:
        ws:                worksheet to draw on.
        row1, col1, row2, col2 (int): 1-indexed range corners, inclusive.
        style (str):        border line style, e.g. "thin" or "double".
        top, bottom, left, right (bool): which edges of the range to draw.
    """
    side = Side(style=style)
    for col in range(col1, col2 + 1):
        if top:
            cell = ws.cell(row=row1, column=col)
            b = cell.border
            cell.border = Border(top=side, bottom=b.bottom, left=b.left, right=b.right)
        if bottom:
            cell = ws.cell(row=row2, column=col)
            b = cell.border
            cell.border = Border(bottom=side, top=b.top, left=b.left, right=b.right)
    for row in range(row1, row2 + 1):
        if left:
            cell = ws.cell(row=row, column=col1)
            b = cell.border
            cell.border = Border(left=side, top=b.top, bottom=b.bottom, right=b.right)
        if right:
            cell = ws.cell(row=row, column=col2)
            b = cell.border
            cell.border = Border(right=side, top=b.top, bottom=b.bottom, left=b.left)


def _write_config_param_block(ws, config_id, variants, show_rl_row, label_col, show_byzantine_row=False):
    """
    Write one config's title + Clients/Epsilon/Rl/Byzantine fraction/TopK
    header rows at label_col, reproducing the hand-built layout in
    tmp/template_excel_output.xlsx (tmp/template_excel_output_third_table.xlsx
    for show_byzantine_row=True).

    Shared by every per-config Excel table this module builds (see
    export_excel_report()) so the title/parameter-row scaffold isn't
    duplicated across them. Draws the title (bold, merged, centered), the
    Clients / Epsilon / TopK rows (plus Rl if show_rl_row, plus Byzantine
    fraction -- between Rl and TopK -- if show_byzantine_row) giving that
    column's value for whichever of those parameters this config actually
    uses ("-" for a mechanism that's off for this config, e.g. Epsilon when
    use_dp=False -- not omitted, so every table has the same row layout and
    lines up across the sheet), and the borders down through the TopK row
    (medium outer box, thin label-column divider, double bottom separator
    under TopK). The caller is responsible for its own rows below the
    returned TopK row, and for extending the outer box/divider down to its
    own last row (same _box_border() calls, starting at topk_row + 1).

    Args:
        ws:                    worksheet to draw on.
        config_id (int):       which config this block is for.
        variants (list[dict]): this config's result JSONs, already sorted
                               (e.g. by variant_sort_key()).
        show_rl_row (bool):    whether to include the Rl row.
        label_col (int):       1-indexed column this block starts at.
        show_byzantine_row (bool): whether to include the Byzantine fraction
                                   row (between Rl and TopK).

    Returns:
        tuple: (last_col, topk_row) -- last_col is this block's last data
              column (label_col + len(variants)); topk_row is the row index
              the caller's own rows should start right after.
    """
    TITLE_ROW, CLIENTS_ROW, EPSILON_ROW = 1, 2, 3
    next_row = 4
    RL_ROW = None
    if show_rl_row:
        RL_ROW, next_row = next_row, next_row + 1
    BYZANTINE_ROW = None
    if show_byzantine_row:
        BYZANTINE_ROW, next_row = next_row, next_row + 1
    TOPK_ROW = next_row
    last_col = label_col + len(variants)

    title_cell = ws.cell(row=TITLE_ROW, column=label_col, value=replace_config_with_label(str(config_id)))
    ws.merge_cells(start_row=TITLE_ROW, start_column=label_col, end_row=TITLE_ROW, end_column=last_col)
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=CLIENTS_ROW, column=label_col, value="Clients")
    ws.cell(row=EPSILON_ROW, column=label_col, value="Epsilon")
    if show_rl_row:
        ws.cell(row=RL_ROW, column=label_col, value="Rl")
    if show_byzantine_row:
        ws.cell(row=BYZANTINE_ROW, column=label_col, value="Byzantine fraction")
    ws.cell(row=TOPK_ROW, column=label_col, value="TopK")

    CENTER = Alignment(horizontal="center")
    for i, result in enumerate(variants):
        config = result["config"]
        data_col = label_col + 1 + i

        ws.cell(row=CLIENTS_ROW, column=data_col, value=config["num_clients"]).alignment = CENTER
        ws.cell(row=EPSILON_ROW, column=data_col,
                value=config["epsilon"] if config["use_dp"] else "-").alignment = CENTER
        if show_rl_row:
            rl = config.get("num_client_iterations_per_round")
            ws.cell(row=RL_ROW, column=data_col, value=rl if rl is not None else "-").alignment = CENTER
        if show_byzantine_row:
            byzantine_fraction = config["num_byzantine"] / config["num_clients"] if config["num_clients"] else 0.0
            ws.cell(row=BYZANTINE_ROW, column=data_col, value=round(byzantine_fraction, 3)).alignment = CENTER
        ws.cell(row=TOPK_ROW, column=data_col,
                value=config["topk_ratio"] if config["use_topk"] else "-").alignment = CENTER

    # Outer box top/left/right down through TOPK_ROW (medium) and title's own bottom separator
    # (medium) -- the caller extends the box/divider down to its own last row once it knows how
    # many rows it needs, and TopK's own bottom gets the double separator instead of the box's
    # medium (drawn last so it wins).
    _box_border(ws, TITLE_ROW, label_col, TOPK_ROW, last_col, style="medium", bottom=False)
    _box_border(ws, TITLE_ROW, label_col, TITLE_ROW, last_col, style="medium", top=False, left=False, right=False)
    _box_border(ws, CLIENTS_ROW, label_col, TOPK_ROW, label_col, style="thin", top=False, bottom=False, left=False)
    _box_border(ws, TOPK_ROW, label_col, TOPK_ROW, last_col, style="double", top=False, left=False, right=False)

    return last_col, TOPK_ROW


def _write_accuracy_sheet(ws, groups, show_rl_row) -> None:
    """
    Build the per-config accuracy-by-round tables on ws, side by side.

    One table per config, left to right in config_id order, via
    _write_config_param_block(); below that, a bold Round/Accuracy header
    row and one data row per round. Variants within a config are ordered by
    variant_sort_key() (TopK outermost, then epsilon, then Rl, then Clients
    innermost), matching bar_accuracy's ordering.

    Args:
        ws:              worksheet to draw on.
        groups (dict):   config_id -> list of that config's result JSONs.
        show_rl_row (bool): whether to include the Rl row (see _write_config_param_block()).
    """
    col = 1
    for config_id in sorted(groups):
        variants = sorted(groups[config_id], key=variant_sort_key)
        label_col = col
        last_col, topk_row = _write_config_param_block(ws, config_id, variants, show_rl_row, label_col)

        header_row = topk_row + 1
        header_cell = ws.cell(row=header_row, column=label_col, value="Round")
        header_cell.font = Font(bold=True)

        first_data_row = header_row + 1
        max_round = 0
        for i, result in enumerate(variants):
            data_col = label_col + 1 + i
            accuracy_header_cell = ws.cell(row=header_row, column=data_col, value="Accuracy")
            accuracy_header_cell.font = Font(bold=True)

            rounds, accuracies = extract_metric(result, "accuracy")
            for r, accuracy in zip(rounds, accuracies):
                ws.cell(row=first_data_row + r - 1, column=data_col, value=accuracy)
            max_round = max(max_round, max(rounds, default=0))

        for r in range(1, max_round + 1):
            ws.cell(row=first_data_row + r - 1, column=label_col, value=r)

        last_row = first_data_row + max_round - 1
        _box_border(ws, header_row, label_col, last_row, last_col, style="medium", top=False)
        _box_border(ws, header_row, label_col, last_row, label_col, style="thin", top=False, bottom=False, left=False)
        _box_border(ws, header_row, label_col, header_row, last_col, style="thin", top=False, left=False, right=False)

        col = last_col + 1


def _write_elapsed_time_sheet(ws, groups, show_rl_row) -> None:
    """
    Build the per-config elapsed-time tables on ws, side by side.

    One table per config, left to right in config_id order, via
    _write_config_param_block(); below that, two summary rows per variant --
    average time per round (elapsed_seconds / num_rounds) and total
    wall-clock time (elapsed_seconds) -- instead of accuracy's per-round
    breakdown, since wall-clock time is a single number per run, not a
    per-round series.

    Args:
        ws:              worksheet to draw on.
        groups (dict):   config_id -> list of that config's result JSONs.
        show_rl_row (bool): whether to include the Rl row (see _write_config_param_block()).
    """
    col = 1
    for config_id in sorted(groups):
        variants = sorted(groups[config_id], key=variant_sort_key)
        label_col = col
        last_col, topk_row = _write_config_param_block(ws, config_id, variants, show_rl_row, label_col)

        avg_row = topk_row + 1
        total_row = topk_row + 2
        ws.cell(row=avg_row, column=label_col, value="Avg s/round")
        ws.cell(row=total_row, column=label_col, value="Total elapsed (s)")

        CENTER = Alignment(horizontal="center")
        for i, result in enumerate(variants):
            data_col = label_col + 1 + i
            elapsed = result["results"].get("elapsed_seconds")
            num_rounds = result["config"]["num_rounds"]
            avg_per_round = elapsed / num_rounds if elapsed is not None and num_rounds else None

            ws.cell(row=avg_row, column=data_col,
                    value=round(avg_per_round, 3) if avg_per_round is not None else None).alignment = CENTER
            ws.cell(row=total_row, column=data_col,
                    value=round(elapsed, 1) if elapsed is not None else None).alignment = CENTER

        _box_border(ws, avg_row, label_col, total_row, last_col, style="medium", top=False)
        _box_border(ws, avg_row, label_col, total_row, label_col, style="thin", top=False, bottom=False, left=False)

        col = last_col + 1


def _zero_byzantine_base_accuracy_by_clients(results: list[dict]) -> dict[int, float]:
    """
    Map num_clients -> final-round accuracy of the no-attack BASE run
    (config_id=1, num_byzantine=0) at that client count.

    Used by _write_summary_sheet() for the "Delta accuracy (gap to
    no-attack BASE)" row. A results folder swept only under attack (every
    config's num_byzantine > 0, e.g. run_configurations.py's default
    SHARED_PARAMS) has no entries here at all -- callers must treat a
    missing client count as "no baseline available", not as a zero delta.

    Args:
        results (list[dict]): loaded result JSONs (see load_results()).

    Returns:
        dict[int, float]: num_clients -> zero-Byzantine BASE final accuracy.
    """
    base_accuracy_by_clients = {}
    for result in results:
        config = result["config"]
        if config["config_id"] != 1 or config["num_byzantine"] != 0:
            continue
        _, accuracies = extract_metric(result, "accuracy")
        if accuracies:
            base_accuracy_by_clients[config["num_clients"]] = accuracies[-1]
    return base_accuracy_by_clients


def _write_summary_sheet(ws, groups, show_rl_row, base_accuracy_by_clients) -> None:
    """
    Build the per-config summary tables on ws, side by side.

    One table per config, left to right in config_id order, via
    _write_config_param_block() (with the Byzantine fraction row shown);
    below that, one row per variant for each of:
      - Rounds: config["num_rounds"].
      - Bytes per client: mean over rounds of that round's total
        update_bytes / num_clients ("bytes per client per round") --
        "-" for result files that predate the update_bytes metric.
      - Delta accuracy: base_accuracy_by_clients's no-attack BASE final
        accuracy at the same client count, minus this variant's final
        accuracy -- "N/A" if base_accuracy_by_clients has no entry for that
        client count (see _zero_byzantine_base_accuracy_by_clients()).
      - Final Accuracy: last round's accuracy.
    Reproduces the hand-built layout in
    tmp/template_excel_output_third_table.xlsx.

    Args:
        ws:              worksheet to draw on.
        groups (dict):   config_id -> list of that config's result JSONs.
        show_rl_row (bool): whether to include the Rl row (see _write_config_param_block()).
        base_accuracy_by_clients (dict[int, float]): num_clients -> no-attack
            BASE final accuracy, see _zero_byzantine_base_accuracy_by_clients().
    """
    CENTER = Alignment(horizontal="center")
    col = 1
    for config_id in sorted(groups):
        variants = sorted(groups[config_id], key=variant_sort_key)
        label_col = col
        last_col, topk_row = _write_config_param_block(
            ws, config_id, variants, show_rl_row, label_col, show_byzantine_row=True)

        rounds_row = topk_row + 1
        bytes_row = topk_row + 2
        delta_row = topk_row + 3
        final_row = topk_row + 4
        ws.cell(row=rounds_row, column=label_col, value="Rounds")
        ws.cell(row=bytes_row, column=label_col, value="Bytes per client")
        ws.cell(row=delta_row, column=label_col, value="Delta accuracy (gap to no-attack BASE)")
        ws.cell(row=final_row, column=label_col, value="Final Accuracy")

        for i, result in enumerate(variants):
            data_col = label_col + 1 + i
            config = result["config"]

            ws.cell(row=rounds_row, column=data_col, value=config["num_rounds"]).alignment = CENTER

            per_client_bytes = [
                entry["update_bytes"] / config["num_clients"]
                for entry in result["results"]["per_round"]
                if entry.get("update_bytes") is not None
            ]
            avg_bytes = sum(per_client_bytes) / len(per_client_bytes) if per_client_bytes else None
            ws.cell(row=bytes_row, column=data_col,
                    value=round(avg_bytes, 1) if avg_bytes is not None else "-").alignment = CENTER

            _, accuracies = extract_metric(result, "accuracy")
            final_accuracy = accuracies[-1] if accuracies else None
            base_accuracy = base_accuracy_by_clients.get(config["num_clients"])
            if base_accuracy is None or final_accuracy is None:
                delta_value = "N/A"
                print("Could not find clean baserun folder. Skipping delta accuracy...")
            else:
                delta_value = round(base_accuracy - final_accuracy, 4)
            ws.cell(row=delta_row, column=data_col, value=delta_value).alignment = CENTER

            ws.cell(row=final_row, column=data_col,
                    value=round(final_accuracy, 4) if final_accuracy is not None else "-").alignment = CENTER

        _box_border(ws, rounds_row, label_col, final_row, last_col, style="medium", top=False)
        _box_border(ws, rounds_row, label_col, final_row, label_col, style="thin", top=False, bottom=False, left=False)

        col = last_col + 1


def _collect_trust_scores(result: dict) -> tuple[list[float], list[float]]:
    """
    Pool every (client, round) trust-score observation for one result,
    split into honest vs malicious.

    Reads results.malicious_clients (a {client_id: is_malicious} dict) and
    each round's trust_scores (a {client_id: score} dict, see
    mechanisms/robust_aggregation.py's FLTrustStrategy.aggregate_fit()).
    Pooling across all rounds (rather than e.g. only the final round)
    mirrors the aggregate honest/malicious comparison in src/README.md's
    "FLTrust trust-score decay investigation" section.

    Args:
        result (dict): loaded result JSON.

    Returns:
        tuple: (honest_scores, malicious_scores) -- flat lists of floats.
              Both empty if this variant has no trust_scores at all (e.g.
              use_fltrust=False).
    """
    malicious_clients = result["results"].get("malicious_clients", {})
    honest_scores = []
    malicious_scores = []
    for entry in result["results"]["per_round"]:
        trust_scores = entry.get("trust_scores")
        if not trust_scores:
            continue
        for client_id, score in trust_scores.items():
            if malicious_clients.get(client_id, False):
                malicious_scores.append(score)
            else:
                honest_scores.append(score)
    return honest_scores, malicious_scores


def _write_trust_per_client_sheet(ws, groups, show_rl_row) -> None:
    """
    Build the per-config trust-per-client tables on ws, side by side.

    One table per config, left to right in config_id order, via
    _write_config_param_block() (with the Byzantine fraction row shown);
    below that, four rows per variant -- mean and std of the honest
    clients' trust scores, then mean and std of the malicious clients'
    trust scores (see _collect_trust_scores()), each pooled across every
    round. "-" for configs with no trust scores at all (use_fltrust=False).
    Reproduces the hand-built layout in
    tmp/template_excel_output_forth_table.xlsx.

    Args:
        ws:              worksheet to draw on.
        groups (dict):   config_id -> list of that config's result JSONs.
        show_rl_row (bool): whether to include the Rl row (see _write_config_param_block()).
    """
    CENTER = Alignment(horizontal="center")
    col = 1
    for config_id in sorted(groups):
        variants = sorted(groups[config_id], key=variant_sort_key)
        label_col = col
        last_col, topk_row = _write_config_param_block(
            ws, config_id, variants, show_rl_row, label_col, show_byzantine_row=True)

        honest_mean_row = topk_row + 1
        honest_std_row = topk_row + 2
        malicious_mean_row = topk_row + 3
        malicious_std_row = topk_row + 4
        ws.cell(row=honest_mean_row, column=label_col, value="TpC mean (honest)")
        ws.cell(row=honest_std_row, column=label_col, value="TpC std (honest)")
        ws.cell(row=malicious_mean_row, column=label_col, value="TpC mean (malicious)")
        ws.cell(row=malicious_std_row, column=label_col, value="TpC std (malicious)")

        for i, result in enumerate(variants):
            data_col = label_col + 1 + i
            honest_scores, malicious_scores = _collect_trust_scores(result)

            ws.cell(row=honest_mean_row, column=data_col,
                    value=round(float(np.mean(honest_scores)), 4) if honest_scores else "-").alignment = CENTER
            ws.cell(row=honest_std_row, column=data_col,
                    value=round(float(np.std(honest_scores)), 4) if honest_scores else "-").alignment = CENTER
            ws.cell(row=malicious_mean_row, column=data_col,
                    value=round(float(np.mean(malicious_scores)), 4) if malicious_scores else "-").alignment = CENTER
            ws.cell(row=malicious_std_row, column=data_col,
                    value=round(float(np.std(malicious_scores)), 4) if malicious_scores else "-").alignment = CENTER

        _box_border(ws, honest_mean_row, label_col, malicious_std_row, last_col, style="medium", top=False)
        _box_border(ws, honest_mean_row, label_col, malicious_std_row, label_col, style="thin", top=False, bottom=False, left=False)

        col = last_col + 1


def export_excel_report(folder: str) -> None:
    """
    Export the full Excel report for a results folder: one workbook, four
    sheets, each laying out one bordered table per config side by side --
    "Accuracy by Round" (per-round accuracy, see _write_accuracy_sheet()),
    "Elapsed Time" (avg time per round + total wall-clock time, see
    _write_elapsed_time_sheet()), "Summary" (Rounds/Bytes per client/Delta
    accuracy/Final Accuracy, see _write_summary_sheet()), and "Trust per
    Client" (honest/malicious trust-score mean+std, see
    _write_trust_per_client_sheet()). Reproduces the hand-built layouts in
    tmp/template_excel_output.xlsx, tmp/template_excel_output_third_table.xlsx,
    and tmp/template_excel_output_forth_table.xlsx for the shared
    title/Clients/Epsilon/Rl/(Byzantine fraction)/TopK header block
    (_write_config_param_block()).

    The Rl row (shared by all four sheets) is only added at all if at least
    one result in folder actually has num_client_iterations_per_round set --
    older result files predate that field entirely, and it's meaningless
    clutter (an all-"-" row) for a folder where nothing set it. Reads every
    JSON file present in folder via load_results(), so this works for any
    number of config runs/variants, same as every other visualize_* function.

    The Summary sheet's "Delta accuracy" row needs a no-attack (num_byzantine=0)
    BASE run to compare against. Since folder's own JSON files are typically
    swept entirely under attack (see run_configurations.py's SHARED_PARAMS),
    that baseline usually doesn't live among them -- it's read from
    folder/clean_base_run/ instead (a subfolder holding a copy of a separate
    no-attack run's result JSONs, see e.g. results/full-grid-run-combined/
    clean_base_run/), if present. Falls back to folder's own results (covers
    folders where the no-attack BASE run *is* one of the swept variants), and
    to "N/A" for any client count found in neither place.

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = load_results(folder)
    show_rl_row = any(r["config"].get("num_client_iterations_per_round") is not None for r in results)

    groups = defaultdict(list)
    for result in results:
        groups[result["config"]["config_id"]].append(result)

    clean_base_run_results = load_results_if_present(os.path.join(folder, "clean_base_run"))
    base_accuracy_by_clients = _zero_byzantine_base_accuracy_by_clients(results + clean_base_run_results)

    wb = Workbook()
    accuracy_ws = wb.active
    accuracy_ws.title = "Accuracy by Round"
    _write_accuracy_sheet(accuracy_ws, groups, show_rl_row)

    elapsed_ws = wb.create_sheet("Elapsed Time")
    _write_elapsed_time_sheet(elapsed_ws, groups, show_rl_row)

    summary_ws = wb.create_sheet("Summary")
    _write_summary_sheet(summary_ws, groups, show_rl_row, base_accuracy_by_clients)

    trust_ws = wb.create_sheet("Trust per Client")
    _write_trust_per_client_sheet(trust_ws, groups, show_rl_row)

    output_path = os.path.join(folder, "accuracy_by_round.xlsx")
    wb.save(_io_path(output_path))
    print(f"Saved to: {output_path}")


def get_base_accuracy_and_max_epsilon(results):
    """
    Find base accuracy and max epsilon in results.

    Args:
        results (list): A list of results dicts

    Returns:
        tuple: (base_accuracy, max_epsilon) - Both values
    """
    base_accuracy = None
    for result in results:
        if result["config"]["config_id"] != 1:
            continue
        base_accuracy = result["results"]["per_round"][-1].get("accuracy")
        if not base_accuracy:
            continue
        break
    if not base_accuracy:
        print("Warning: Could not find base_accuracy! Most likely cause is that there is no config-1 results file. Falling back to best accuracy as baseline ...")
        base_accuracy = max(
            r["results"]["per_round"][-1].get("accuracy", 0)
            for r in results
        )

    # find max epsilon to normalize
    max_epsilon = float('-inf')
    for result in results:
        if not result["config"]["use_dp"]:
            continue
        this_epsilon = result["results"]["per_round"][-1].get("epsilon")
        if this_epsilon and this_epsilon > max_epsilon:
            max_epsilon = this_epsilon
    if max_epsilon < 0:
        print(f"Could not find a valid epsilon across all results in the specified folder. Found maximum epsilon was epsilon={max_epsilon}.")

    return base_accuracy, max_epsilon

def visualize_radar_chart(folder: str) -> None:
    """
    Radar/spider chart showing privacy-robustness-performance tradeoff.

    Each config variant gets one polygon on the radar.
    Three axes: Privacy (1 - normalized_epsilon), Robustness (accuracy_under_attack),
    Performance (final_accuracy_no_attack baseline comparison).

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = load_results(folder)

    _, max_epsilon = get_base_accuracy_and_max_epsilon(results)

    normalized_result_scores = defaultdict(dict)
    PRIVACY_KEY = "privacy_score"
    ROBUSTNESS_KEY = "robustness_score"
    EFFICIENCY_KEY = "efficiency_score"
    for result in results:
        config = result["config"]
        final_accuracy = result["results"]["per_round"][-1].get("accuracy")
        
        # privacy
        if not config["use_dp"]:
            normalized_result_scores[result["_filename"]][PRIVACY_KEY] = 0.0         # No privacy at all
        else:
            normalized_result_scores[result["_filename"]][PRIVACY_KEY] = max(1 - (config["epsilon"] / max_epsilon), 0) # The max is due to currently possible negative epsilon values
        
        # robustness
        normalized_result_scores[result["_filename"]][ROBUSTNESS_KEY] = final_accuracy

        # efficiency
        if not config["use_topk"]:
            normalized_result_scores[result["_filename"]][EFFICIENCY_KEY] = 0.0
        else:
            normalized_result_scores[result["_filename"]][EFFICIENCY_KEY] = 1 - config["topk_ratio"] 
            


    variants = [(name, values) for name, values in normalized_result_scores.items()]

    axes_labels = ["Privacy", "Robustness", "Efficiency"]
    num_axes    = len(axes_labels)
    # Privacy at top (90°), Robustness bottom-left (210°), Efficiency bottom-right (330°)
    angles = [np.pi/2, np.pi/2 + 2*np.pi/3, np.pi/2 + 4*np.pi/3]
    angles += angles[:1]    # Close the shape

    fig, ax = plt.subplots(figsize=(8, 8),
                           subplot_kw=dict(polar=True))
    ax.set_title("Privacy - Robustness - Efficiency Trilemma",
                 fontsize=13, pad=20)

    colors = plt.cm.tab10.colors
    for i, (label, scores) in enumerate(variants):
        values = [
            scores[PRIVACY_KEY],
            scores[ROBUSTNESS_KEY],
            scores[EFFICIENCY_KEY],
        ]
        values += values[:1]   # close the polygon
        color   = colors[i % len(colors)]
        ax.plot(angles, values, color=color, linewidth=1.5, label=label)
        ax.fill(angles, values, color=color, alpha=0.1)

    ax.set_thetagrids(
        [a * 180 / np.pi for a in angles[:-1]],
        axes_labels,
        fontsize=11,
    )
    ax.set_ylim(0, 1)
    ax.set_yticks([0.25, 0.5, 0.75, 1.0])
    ax.set_yticklabels(["0.25", "0.50", "0.75", "1.00"], fontsize=7)
    ax.grid(True, alpha=0.3)

    ax.legend(
        fontsize=6,
        loc="upper center",
        bbox_to_anchor=(0.5, -0.08),
        ncol=2,
        framealpha=0.8,
    )

    plt.tight_layout()
    output_path = os.path.join(folder, "radar_trilemma.png")
    plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
    print(f"Saved to: {output_path}")
    plt.close()


def visualize_radar_chart_per_config(folder: str) -> None:
    """
    Radar/spider chart showing privacy-robustness-performance tradeoff.

    One Subplot per config.

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = load_results(folder)

    _, max_epsilon = get_base_accuracy_and_max_epsilon(results)

    PRIVACY_KEY = "privacy_score"
    ROBUSTNESS_KEY = "robustness_score"
    EFFICIENCY_KEY = "efficiency_score"
    config_groups = defaultdict(list)
    for result in results:
        config = result["config"]
        final_accuracy = result["results"]["per_round"][-1].get("accuracy")
        
        # privacy
        if not config["use_dp"]:
            privacy = 0.0         # No privacy at all
        else:
            privacy = max(1 - (config["epsilon"] / max_epsilon), 0) # DP - The max is due to currently possible negative epsilon values
        
        # robustness
        robustness = final_accuracy                 # FLTrust

        # efficiency
        if not config["use_topk"]:
            efficiency = 0.0
        else:
            efficiency = 1 - config["topk_ratio"]                   # TopK

        config_groups[config["config_id"]].append((
            make_label(result["_filename"]),
            {PRIVACY_KEY: privacy, ROBUSTNESS_KEY: robustness, EFFICIENCY_KEY: efficiency}
        ))
            

    axes_labels    = ["Privacy", "Robustness", "Efficiency"]

    angles  = [np.pi/2, np.pi/2 + 2*np.pi/3, np.pi/2 + 4*np.pi/3]
    angles += angles[:1]

    fig, axes = plt.subplots(2, 4, figsize=(20, 10),
                             subplot_kw=dict(polar=True))
    fig.suptitle("Privacy - Robustness - Efficiency Trilemma per Config",
                 fontsize=14)
    axes = axes.flatten()

    colors = plt.cm.tab10.colors

    for idx, config_id in enumerate(sorted(config_groups.keys())):
        ax      = axes[idx]
        variants = config_groups[config_id]

        ax.set_title(replace_config_with_label(str(config_id)), fontsize=10, pad=10)

        for i, (label, scores) in enumerate(variants):
            values  = [scores[PRIVACY_KEY],
                       scores[ROBUSTNESS_KEY],
                       scores[EFFICIENCY_KEY]]
            values += values[:1]
            color   = colors[i % len(colors)]
            ax.plot(angles, values, color=color,
                    linewidth=1.5, label=label)
            ax.fill(angles, values, color=color, alpha=0.15)

        ax.set_thetagrids(
            [a * 180 / np.pi for a in angles[:-1]],
            axes_labels, fontsize=8,
        )
        ax.set_ylim(0, 1)
        ax.set_yticks([0.25, 0.5, 0.75, 1.0])
        ax.set_yticklabels(["0.25", "0.50", "0.75", "1.00"], fontsize=5)
        ax.grid(True, alpha=0.3)
        ax.legend(fontsize=5, loc="upper center",
                  bbox_to_anchor=(0.5, -0.12),
                  ncol=1, framealpha=0.8)

    # hide unused subplots if fewer than 8 configs
    for idx in range(len(config_groups), len(axes)):
        axes[idx].set_visible(False)

    plt.tight_layout()
    output_path = os.path.join(folder, "radar_per_config.png")
    plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
    print(f"Saved to: {output_path}")
    plt.close()


def _descriptive_variant_filename(filename: str, ext: str) -> str:
    """
    Build a descriptive, filesystem-safe filename for one result variant.

    Reuses run_configurations.py's make_filename() naming (dataset/config/
    dp/fltrust/topk/rounds/clients/byzantine/attack/...), stripping its
    leading run_timestamp -- unique already, but not descriptive of the
    variant itself, and the containing run folder (which may combine
    several timestamped runs, e.g. results/full-grid-run-combined/) already
    scopes it. Anchored on the "dataset-" token (always the first
    descriptive part make_filename() writes) rather than a fixed token
    count, so it doesn't assume a particular timestamp format.

    Args:
        filename (str): original result JSON filename (result["_filename"]).
        ext (str):       output file extension, without the dot, e.g. "png".

    Returns:
        str: e.g. "dataset-mnist_config-3_dp-False_fltrust-True_topk-False_
             rounds-500_clients-10_byzantine-2_attack-label_flip_scale-2.0_Rl-10.png".
    """
    parts = filename.replace(".json", "").split("_")
    start = next((i for i, part in enumerate(parts) if part.startswith("dataset-")), 0)
    return "_".join(parts[start:]) + f".{ext}"


def visualize_confusion_matrices(folder: str) -> None:
    """
    One confusion-matrix heatmap PNG per result JSON.

    Previously rendered a single combined PNG with one subplot per config,
    each showing only that config's best-accuracy variant via
    group_by_config_id() -- every other variant's confusion matrix was
    silently discarded (see TODO.md's "What to do next"). Now saves one
    file per result JSON instead, under
    confusion_matrices/config_<id>/<descriptive_filename>.png -- a config
    subfolder one level deeper than f1_tables/'s layout, since every
    variant now gets its own file rather than one shared per config.
    Results with no confusion_matrix (e.g. very old files) are skipped.

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = load_results(folder)
    num_classes  = get_dataset_spec(results[0]["config"].get("dataset", "mnist")).num_classes
    digit_labels = [str(i) for i in range(num_classes)]

    saved = 0
    for result in results:
        matrix = result["results"].get("confusion_matrix")
        if matrix is None:
            continue

        config_id = result["config"]["config_id"]
        label     = make_label(result["_filename"])
        mat       = np.array(matrix)

        # normalize by row (true label) to show percentages
        row_sums = mat.sum(axis=1, keepdims=True)
        mat_norm = np.where(row_sums > 0, mat / row_sums, 0)

        fig, ax = plt.subplots(figsize=(5, 5))
        im = ax.imshow(mat_norm, interpolation="nearest", cmap="Blues", vmin=0, vmax=1)
        ax.set_title(f"{replace_config_with_label(str(config_id))} - {label}", fontsize=9)
        ax.set_xlabel("Predicted")
        ax.set_ylabel("True")
        ax.set_xticks(range(num_classes))
        ax.set_yticks(range(num_classes))
        ax.set_xticklabels(digit_labels, fontsize=7)
        ax.set_yticklabels(digit_labels, fontsize=7)

        # annotate cells with percentage
        for i in range(num_classes):
            for j in range(num_classes):
                ax.text(j, i, f"{mat_norm[i,j]:.0%}",
                        ha="center", va="center",
                        fontsize=6,
                        color="white" if mat_norm[i,j] > 0.5 else "black")

        plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
        plt.tight_layout()

        config_folder = os.path.join(folder, "confusion_matrices", f"config_{config_id}")
        os.makedirs(config_folder, exist_ok=True)
        output_path = os.path.join(config_folder, _descriptive_variant_filename(result["_filename"], "png"))
        plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
        plt.close()
        saved += 1

    print(f"Saved {saved} confusion matrices to: {os.path.join(folder, 'confusion_matrices')}")


def visualize_f1_score(folder: str) -> None:
    """
    One per-class F1/precision/recall bar-chart PNG per result JSON.

    Previously rendered a single combined PNG with one subplot per config,
    each showing only that config's best-accuracy variant via
    group_by_config_id() -- every other variant's scores were silently
    discarded (see TODO.md's "What to do next"). Now saves one file per
    result JSON instead, under f1_charts/config_<id>/<descriptive_filename>.png,
    mirroring visualize_confusion_matrices()'s restructuring. Results with
    no per_class_scores (e.g. very old files) are skipped.

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = load_results(folder)
    num_classes  = get_dataset_spec(results[0]["config"].get("dataset", "mnist")).num_classes
    digit_labels = [str(i) for i in range(num_classes)]

    metrics = ["precision", "recall", "f1"]
    colors  = ["#4C72B0", "#55A868", "#C44E52"]
    x       = np.arange(num_classes)
    width   = 0.25

    saved = 0
    for result in results:
        scores = result["results"].get("per_class_scores")
        if scores is None:
            continue

        config_id = result["config"]["config_id"]
        label     = make_label(result["_filename"])

        fig, ax = plt.subplots(figsize=(6, 4.5))
        for m_idx, (metric, color) in enumerate(zip(metrics, colors)):
            values = [scores[str(d)][metric] for d in range(num_classes)]
            ax.bar(x + m_idx * width, values, width, label=metric, color=color, alpha=0.85)

        ax.set_title(f"{replace_config_with_label(str(config_id))} - {label}", fontsize=9)
        ax.set_xlabel("Digit class")
        ax.set_ylabel("Score")
        ax.set_xticks(x + width)
        ax.set_xticklabels(digit_labels, fontsize=8)
        ax.set_ylim(0, 1.1)
        ax.grid(axis="y", alpha=0.3)
        ax.legend(fontsize=7, loc="lower right")
        plt.tight_layout()

        config_folder = os.path.join(folder, "f1_charts", f"config_{config_id}")
        os.makedirs(config_folder, exist_ok=True)
        output_path = os.path.join(config_folder, _descriptive_variant_filename(result["_filename"], "png"))
        plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
        plt.close()
        saved += 1

    print(f"Saved {saved} F1 charts to: {os.path.join(folder, 'f1_charts')}")


def visualize_f1_tables(folder: str) -> None:
    """
    One summary table PNG per result JSON showing TP/FP/FN/TN and F1 metrics
    per digit class.

    Previously kept only each config's best-accuracy variant (via
    group_by_config_id()), saving one file per config into a flat
    'f1_tables' subfolder. Now saves one file per result JSON instead,
    under f1_tables/config_<id>/<descriptive_filename>.png -- one folder
    deeper, mirroring visualize_confusion_matrices()'s and
    visualize_f1_score()'s restructuring (see TODO.md's "What to do next").
    Results missing per_class_scores or confusion_matrix (e.g. very old
    files) are skipped.

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = load_results(folder)
    num_classes = get_dataset_spec(results[0]["config"].get("dataset", "mnist")).num_classes

    saved = 0
    for result in results:
        scores = result["results"].get("per_class_scores")
        matrix = result["results"].get("confusion_matrix")
        if scores is None or matrix is None:
            continue

        config_id = result["config"]["config_id"]
        label     = make_label(result["_filename"])

        fig, ax = plt.subplots(figsize=(12, 4))
        test_size = sum(sum(row) for row in matrix)
        ax.set_title(f"{replace_config_with_label(str(config_id))} - {label} (n={test_size:,})", fontsize=10, pad=10)
        ax.axis("off")

        col_labels   = ["Class", "TP", "FP", "FN", "TN", "Precision", "Recall", "F1"]
        table_data   = []

        for i in range(num_classes):
            tp = matrix[i][i]
            fp = sum(matrix[r][i] for r in range(num_classes)) - tp
            fn = sum(matrix[i][c] for c in range(num_classes)) - tp
            tn = sum(
                matrix[r][c]
                for r in range(num_classes)
                for c in range(num_classes)
            ) - tp - fp - fn

            precision = scores[str(i)]["precision"]
            recall    = scores[str(i)]["recall"]
            f1        = scores[str(i)]["f1"]

            table_data.append([
                str(i),
                f"{tp:,}", f"{fp:,}", f"{fn:,}", f"{tn:,}",
                f"{precision:.3f}", f"{recall:.3f}", f"{f1:.3f}",
            ])

        # Sum row
        total_tp = sum(matrix[i][i] for i in range(num_classes))
        total_fp = sum(
            sum(matrix[r][i] for r in range(num_classes)) - matrix[i][i]
            for i in range(num_classes)
        )
        total_fn = sum(
            sum(matrix[i][c] for c in range(num_classes)) - matrix[i][i]
            for i in range(num_classes)
        )
        total_tn = sum(
            sum(matrix[r][c] for r in range(num_classes) for c in range(num_classes))
            - matrix[i][i]
            - (sum(matrix[r][i] for r in range(num_classes)) - matrix[i][i])
            - (sum(matrix[i][c] for c in range(num_classes)) - matrix[i][i])
            for i in range(num_classes)
        )

        # Macro averages for precision, recall, f1
        macro_precision = sum(scores[str(i)]["precision"] for i in range(num_classes)) / num_classes
        macro_recall    = sum(scores[str(i)]["recall"]    for i in range(num_classes)) / num_classes
        macro_f1        = sum(scores[str(i)]["f1"]        for i in range(num_classes)) / num_classes

        table_data.append([
            "Total",
            f"{total_tp:,}", f"{total_fp:,}", f"{total_fn:,}", f"{total_tn:,}",
            f"{macro_precision:.3f}", f"{macro_recall:.3f}", f"{macro_f1:.3f}",
        ])

        table = ax.table(
            cellText=table_data,
            colLabels=col_labels,
            loc="center",
            cellLoc="center",
        )
        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1, 1.5)

        # Color the sum row
        for j in range(len(col_labels)):
            table[num_classes + 1, j].set_facecolor("#D0D0D0")
            table[num_classes + 1, j].set_text_props(fontweight="bold")


        # color the diagonal (correct predictions) green, others light red
        for i in range(num_classes):
            # header row is row 0, data starts at row 1
            for j, col in enumerate(col_labels):
                cell = table[i + 1, j]
                if col == "F1":
                    f1_val = scores[str(i)]["f1"]
                    # green gradient based on F1 score
                    cell.set_facecolor(
                        (1 - f1_val * 0.5, 1, 1 - f1_val * 0.5)
                    )
                elif col == "Class":
                    cell.set_facecolor("#E8E8E8")

        plt.tight_layout()
        config_folder = os.path.join(folder, "f1_tables", f"config_{config_id}")
        os.makedirs(config_folder, exist_ok=True)
        output_path = os.path.join(config_folder, _descriptive_variant_filename(result["_filename"], "png"))
        plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
        plt.close()
        saved += 1

    print(f"Saved {saved} F1 tables to: {os.path.join(folder, 'f1_tables')}")


def visualize_label_distribution(folder: str) -> None:
    """
    Stacked bar chart of label counts in the root set and each client's shard.

    The split is identical across every variant in a run folder (same
    root_dataset_size/num_clients/seed), so this reads the label_distribution
    block from any one result file. Older result files saved before this
    field existed are skipped gracefully.

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = load_results(folder)

    label_distribution = None
    for result in results:
        label_distribution = result["results"].get("label_distribution")
        if label_distribution is not None:
            break

    if label_distribution is None:
        print("No label_distribution field found in any result file -- skipping label distribution plot.")
        return

    os.makedirs(os.path.join(folder, "label_distribution"), exist_ok=True)

    num_classes  = get_dataset_spec(results[0]["config"].get("dataset", "mnist")).num_classes
    digit_labels = [str(i) for i in range(num_classes)]
    colors = plt.cm.tab10.colors

    bar_names = ["Root"] + [f"C{cid}" for cid in sorted(label_distribution["clients"], key=int)]
    bar_data  = [label_distribution["root"]] + [
        label_distribution["clients"][cid] for cid in sorted(label_distribution["clients"], key=int)
    ]

    fig, ax = plt.subplots(figsize=(max(10, len(bar_names) * 0.4), 6))
    ax.set_title("Label Distribution: Root Dataset + Each Client's Train Shard", fontsize=13)
    ax.set_ylabel("Sample count")
    ax.grid(axis="y", alpha=0.3)

    x = np.arange(len(bar_names))
    bottom = np.zeros(len(bar_names))
    for digit, color in zip(digit_labels, colors):
        heights = np.array([counts.get(digit, 0) for counts in bar_data])
        ax.bar(x, heights, bottom=bottom, color=color, width=0.8, label=digit)
        bottom += heights

    ax.set_xticks(x)
    ax.set_xticklabels(bar_names, fontsize=6 if len(bar_names) > 20 else 8, rotation=90)
    ax.legend(title="Digit", fontsize=7, loc="upper center",
              bbox_to_anchor=(0.5, -0.15), ncol=10, framealpha=0.8)

    plt.tight_layout()
    output_path = os.path.join(folder, "label_distribution", "label_distribution.png")
    plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
    print(f"Saved to: {output_path}")
    plt.close()


# TopK ratios to show in visualize_bytes_per_client_topk() -- 0.01/0.1/0.5 match
# run_configurations.py's planned/actual TOPK_VALUES sweeps, 1.0 is the dense/no-TopK
# reference bar.
BYTES_CHART_TOPK_RATIOS = [0.01, 0.1, 0.5, 1.0]


def visualize_bytes_per_client_topk(folder: str) -> None:
    """
    Bar chart of per-client, per-round communication size at different TopK
    keep ratios (BYTES_CHART_TOPK_RATIOS), plus a dense/no-TopK reference bar.

    Deliberately not tied to any experiment's results: per-client bytes only
    depend on the dataset's model size and topk_ratio, never on num_clients,
    DP, FLTrust, seed, etc. (see mechanisms/topk.py). So this doesn't call
    load_results() and doesn't require the folder to contain anything --
    `folder` is used purely as an output location, same convention every
    other plot in this file already follows. If the folder does contain
    result files, the first one's `config.dataset` is used so a CIFAR-10
    run's folder gets CIFAR-10-sized bars; otherwise defaults to "mnist".
    Reuses the actual production math (`topk_sparsify()` + `update_size_bytes()`)
    against a dummy parameter vector instead of recomputing the formula by
    hand, so this chart can't drift out of sync with what those functions
    compute for a real run.

    This calls `update_size_bytes()` the same way a single client's own
    `fit()` does (client.py/mechanisms/attacks.py) -- i.e. every bar is one
    client's per-round bytes. That is *not* the same number as a result
    JSON's per-round `update_bytes` field, which is the sum across all
    participating clients that round (see HistoryStrategyAdapter.aggregate_fit()
    in server.py, and the "communication-size metric" entry in
    src/README.md) and therefore scales with num_clients -- this chart
    deliberately shows the num_clients-independent, single-client figure
    instead.

    Caveat (not drawn on the chart, so it doesn't get lost if the image is
    used standalone): per mechanisms/topk.py's own docstring, the k<1.0 bars
    are the *logical* sparse-encoding cost a real (index, value) wire format
    would need, not what's actually transmitted in this simulation today --
    `fit()` still returns a dense array regardless of topk_ratio, so every
    config's real current bytes-on-wire matches the k=1.0 (dense) bar. See
    src/README.md item 15 for the full writeup.

    Args:
        folder (str): folder to save the chart into. Read from (best-effort
                      dataset detection) if it already contains result files,
                      but not required to.
    """
    dataset = "mnist"
    for filepath in sorted(glob.glob(os.path.join(folder, "*.json"))):
        if os.path.basename(filepath) == "run_summary.json":
            continue
        with open(_io_path(filepath)) as f:
            dataset = json.load(f).get("config", {}).get("dataset", "mnist")
        break

    dataset_spec = get_dataset_spec(dataset)
    num_params = sum(p.numel() for p in dataset_spec.model_fn().parameters())
    dummy_update = np.ones(num_params, dtype=np.float32)   # magnitudes are irrelevant, only count/dtype size matter

    bar_labels, bytes_per_ratio = [], []
    for k in BYTES_CHART_TOPK_RATIOS:
        if k == 1.0:
            bar_labels.append("Dense\n(no TopK)")
            bytes_per_ratio.append(update_size_bytes([dummy_update], use_topk=False))
        else:
            bar_labels.append(f"{k:.0%}")
            sparsified = topk_sparsify(dummy_update, k)
            bytes_per_ratio.append(update_size_bytes(None, use_topk=True, sparsified_update=sparsified))

    kb_per_ratio = [b / 1000 for b in bytes_per_ratio]  # decimal KB, not KiB

    os.makedirs(folder, exist_ok=True)
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.set_title(f"Per-Client Communication Size with TopK ({dataset})", fontsize=13)
    ax.set_ylabel("KB per client per round")
    ax.grid(axis="y", alpha=0.3)
    bars = ax.bar(bar_labels, kb_per_ratio, color=plt.cm.tab10.colors[:len(bar_labels)])
    ax.bar_label(bars, fmt="%.1f KB", fontsize=8)

    plt.tight_layout()
    output_path = os.path.join(folder, "bytes_per_client.png")
    plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
    print(f"Saved to: {output_path}")
    plt.close()


# Fixed colors for the honest/malicious split, reused across every trust
# plot below -- never cycled or reassigned per config/client like the tab10
# colors elsewhere in this file, since honest vs. malicious is a fixed,
# always-two-valued category.
HONEST_COLOR = "#1f77b4"
MALICIOUS_COLOR = "#d62728"


def _avg_trust_per_client(result: dict) -> dict:
    """
    Average one result's per-round trust scores into {client_id: avg_trust}.

    Args:
        result (dict): loaded result JSON (must have use_fltrust=True).

    Returns:
        dict: {client_id (str): average trust score over the whole run}.
    """
    sums, counts = {}, {}
    for entry in result["results"]["per_round"]:
        for client_id, score in entry["trust_scores"].items():
            sums[client_id] = sums.get(client_id, 0.0) + score
            counts[client_id] = counts.get(client_id, 0) + 1
    return {client_id: sums[client_id] / counts[client_id] for client_id in sums}


def visualize_trust_per_client(folder: str) -> None:
    """
    Average trust per client over the whole run, grouped by FLTrust config variant.

    One point per client per variant, colored by whether that client was
    honest or malicious, with a small horizontal spread (not random jitter,
    so the plot is reproducible) so points inside a group don't overlap.
    Non-FLTrust variants never produce trust scores, so they're excluded.
    DP+FLTrust variants are included -- DP noise degrades the cosine-
    similarity trust signal, but seeing that degradation next to plain
    FLTrust's clean separation is itself informative.

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = [r for r in load_results(folder) if r["config"]["use_fltrust"]]
    if not results:
        print("No non-DP FLTrust-enabled result files found -- skipping per-client trust plot.")
        return
    results.sort(key=lambda r: (r["config"]["config_id"], r["_filename"]))

    fig, ax = plt.subplots(figsize=(max(10, len(results) * 1.3), 6))
    ax.set_title("Average Trust per Client over the Run (FLTrust configs only)", fontsize=13)
    ax.set_ylabel("Avg trust score")
    ax.set_ylim(0, 1.05)
    ax.grid(axis="y", alpha=0.3)

    tick_pos, tick_lab = [], []
    for x, result in enumerate(results):
        malicious = result["results"]["malicious_clients"]
        avg_trust = _avg_trust_per_client(result)
        client_ids = sorted(avg_trust, key=lambda c: (malicious.get(c, False), int(c)))

        for i, client_id in enumerate(client_ids):
            spread = (i / max(len(client_ids) - 1, 1) - 0.5) * 0.5   # spread across [-0.25, 0.25]
            color = MALICIOUS_COLOR if malicious.get(client_id, False) else HONEST_COLOR
            ax.scatter(x + spread, avg_trust[client_id], color=color,
                       edgecolor="white", linewidth=0.5, s=35, alpha=0.85, zorder=3)

        honest_vals = [avg_trust[c] for c in client_ids if not malicious.get(c, False)]
        malicious_vals = [avg_trust[c] for c in client_ids if malicious.get(c, False)]
        if honest_vals:
            ax.hlines(np.mean(honest_vals), x - 0.3, x + 0.3, color=HONEST_COLOR, linewidth=2, zorder=4)
        if malicious_vals:
            ax.hlines(np.mean(malicious_vals), x - 0.3, x + 0.3, color=MALICIOUS_COLOR, linewidth=2, zorder=4)

        tick_pos.append(x)
        tick_lab.append(make_label(result["_filename"]))

    ax.set_xticks(tick_pos)
    ax.set_xticklabels(tick_lab, fontsize=7, rotation=45, ha="right")

    legend_handles = [
        plt.Line2D([0], [0], marker="o", color="none", markerfacecolor=HONEST_COLOR, markersize=7, label="Honest client"),
        plt.Line2D([0], [0], marker="o", color="none", markerfacecolor=MALICIOUS_COLOR, markersize=7, label="Malicious client"),
        plt.Line2D([0], [0], color="black", linewidth=2, label="Group mean"),
    ]
    ax.legend(handles=legend_handles, fontsize=8, loc="upper center",
              bbox_to_anchor=(0.5, -0.3), ncol=3, framealpha=0.8)

    plt.tight_layout()
    output_path = os.path.join(folder, "trust_per_client.png")
    plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
    print(f"Saved to: {output_path}")
    plt.close()


def _honest_vs_malicious_by_round(results: list[dict]) -> tuple[list[int], list, list]:
    """
    Pool per-round trust scores across results into honest/malicious round averages.

    Args:
        results (list[dict]): loaded result JSONs (all use_fltrust=True).

    Returns:
        tuple: (rounds, honest_avg, malicious_avg) -- parallel lists, one
              entry per round; averages are None for a round with no scores
              of that kind.
    """
    honest_by_round = defaultdict(list)
    malicious_by_round = defaultdict(list)
    for result in results:
        malicious = result["results"]["malicious_clients"]
        for entry in result["results"]["per_round"]:
            for client_id, score in entry["trust_scores"].items():
                bucket = malicious_by_round if malicious.get(client_id, False) else honest_by_round
                bucket[entry["round"]].append(score)

    rounds = sorted(set(honest_by_round) | set(malicious_by_round))
    honest_avg = [np.mean(honest_by_round[r]) if honest_by_round[r] else None for r in rounds]
    malicious_avg = [np.mean(malicious_by_round[r]) if malicious_by_round[r] else None for r in rounds]
    return rounds, honest_avg, malicious_avg


def visualize_trust_over_rounds(folder: str) -> None:
    """
    Honest vs. malicious average trust score per round, pooled across every
    FLTrust-enabled config variant in the folder, DP+FLTrust included -- DP
    noise degrades the cosine-similarity trust signal, but pooling it in
    alongside plain FLTrust shows that degradation rather than hiding it.

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = [r for r in load_results(folder) if r["config"]["use_fltrust"]]
    if not results:
        print("No FLTrust-enabled result files found -- skipping trust-over-rounds plot.")
        return

    rounds, honest_avg, malicious_avg = _honest_vs_malicious_by_round(results)

    fig, ax = plt.subplots(figsize=(10, 6))
    ax.set_title("Honest vs Malicious Avg Trust per Round (pooled across all FLTrust configs)", fontsize=13)
    ax.set_xlabel("Round")
    ax.set_ylabel("Avg trust score")
    ax.set_ylim(0, 1.05)
    ax.grid(True, alpha=0.3)
    ax.plot(rounds, honest_avg, color=HONEST_COLOR, linewidth=2, label="Honest")
    ax.plot(rounds, malicious_avg, color=MALICIOUS_COLOR, linewidth=2, label="Malicious")
    ax.legend(fontsize=9, loc="best")

    plt.tight_layout()
    output_path = os.path.join(folder, "trust_over_rounds.png")
    plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
    print(f"Saved to: {output_path}")
    plt.close()


def visualize_trust_over_rounds_per_config(folder: str) -> None:
    """
    Honest vs. malicious average trust score per round, one file per
    FLTrust-enabled config family (pooling that config's own variants, e.g.
    its different topk_ratio values) -- so TopK's effect on trust isn't
    averaged away against plain FLTrust. DP+FLTrust configs are included --
    DP noise degrades the cosine-similarity trust signal, but seeing that
    degradation next to plain FLTrust's clean separation is itself
    informative. Saves one PNG per config into a subfolder, same pattern as
    `visualize_f1_tables()`/`visualize_lines_per_config()`.

    Args:
        folder (str): path to folder containing result JSON files.
    """
    results = [r for r in load_results(folder) if r["config"]["use_fltrust"]]
    if not results:
        print("No FLTrust-enabled result files found -- skipping per-config trust-over-rounds plot.")
        return

    config_groups = defaultdict(list)
    for result in results:
        config_groups[result["config"]["config_id"]].append(result)

    subfolder_name = "trust_over_rounds_per_config"
    os.makedirs(os.path.join(folder, subfolder_name), exist_ok=True)

    for config_id, group_results in sorted(config_groups.items()):
        rounds, honest_avg, malicious_avg = _honest_vs_malicious_by_round(group_results)

        fig, ax = plt.subplots(figsize=(8, 6))
        ax.set_title(f"{replace_config_with_label(str(config_id))} - Honest vs Malicious Avg Trust per Round", fontsize=12)
        ax.set_xlabel("Round")
        ax.set_ylabel("Avg trust score")
        ax.set_ylim(0, 1.05)
        ax.grid(True, alpha=0.3)
        ax.plot(rounds, honest_avg, color=HONEST_COLOR, linewidth=1.5, label="Honest")
        ax.plot(rounds, malicious_avg, color=MALICIOUS_COLOR, linewidth=1.5, label="Malicious")
        ax.legend(fontsize=9, loc="best")

        plt.tight_layout()
        output_path = os.path.join(folder, subfolder_name, f"trust_over_rounds_config_{config_id}.png")
        plt.savefig(_io_path(output_path), dpi=150, bbox_inches="tight")
        print(f"Saved to: {output_path}")
        plt.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Visualize FL trilemma results from a results folder.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "folder",
        type=str,
        help="path to folder containing result JSON files",
    )
    parser.add_argument("--plot", type=str,
                        choices=["all", "lines", "bar", "radar", "radar_per_conf", "lines_per_conf",
                                 "confusion", "f1_chart", "f1_table_per_conf", "label_dist",
                                 "trust_per_client", "trust_rounds", "trust_rounds_per_conf", "excel",
                                 "duration_per_conf", "bytes_bar"],
                        default="all",
                        help="which plot to generate (default: all). 'bytes_bar' is not included in "
                             "'all' since it's a model-size fact, not a per-run result")
    args = parser.parse_args(args=None if len(sys.argv) > 1 else ["--help"])

    if args.plot in ("all", "lines"):
        visualize_overview(args.folder)
    if args.plot in ("all", "bar"):
        visualize_bar_chart_per_config(args.folder)
    if args.plot in ("all", "radar"):
        visualize_radar_chart(args.folder)
    if args.plot in ("all", "radar_per_conf"):
        visualize_radar_chart_per_config(args.folder)
    if args.plot in ("all", "lines_per_conf"):
        visualize_lines_per_config(args.folder)
    if args.plot in ("all", "confusion"):
        visualize_confusion_matrices(args.folder)
    if args.plot in ("all", "f1_chart"):
        visualize_f1_score(args.folder)
    if args.plot in ("all", "label_dist"):
        visualize_label_distribution(args.folder)
    if args.plot in ("all", "f1_table_per_conf"):
        visualize_f1_tables(args.folder)
    if args.plot in ("all", "trust_per_client"):
        visualize_trust_per_client(args.folder)
    if args.plot in ("all", "trust_rounds"):
        visualize_trust_over_rounds(args.folder)
    if args.plot in ("all", "trust_rounds_per_conf"):
        visualize_trust_over_rounds_per_config(args.folder)
    if args.plot in ("all", "excel"):
        export_excel_report(args.folder)
    if args.plot in ("all", "duration_per_conf"):
        visualize_duration_per_config(args.folder)
    if args.plot == "bytes_bar":
        visualize_bytes_per_client_topk(args.folder)